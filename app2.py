import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt 
import seaborn as sns
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def graficos():
    st.subheader("Data Visualization Dashboard")
    st.markdown("### Explorando diferentes bibliotecas de visualización en Python")

    with st.expander("Introducción", expanded=True):
        st.markdown("""
        Esta aplicación demuestra el uso de diferentes bibliotecas de visualización en Python:
        * **Matplotlib**: Biblioteca base para visualización
        * **Seaborn**: Visualizaciones estadísticas de alto nivel
        * **Plotly**: Gráficos interactivos
        * **Streamlit**: Framework para aplicaciones de datos
        """)

    try:
        #base_df = pd.read_excel("datos/BASE.xlsx", sheet_name="BASE")
        base_df = pd.read_excel("BASE.xlsx", sheet_name="BASE")
        #base_df = load_workbook("datos/BASE.xlsx")
        #base_df = base_df["BASE"] 
        dato_df = pd.read_csv('GOOGLE_BASE_WINE_CHEESE_PROJECT - DATASET.csv')
        st.success(" Datos cargados exitosamente")

        st.header("Visualizaciones con Matplotlib")

        with st.container():
            col1, col2 = st.columns(2)

            with col1:
                st.subheader("Gráfico de Dispersión")
                fig, ax = plt.subplots(figsize=(8, 6))
                TipoVinho = base_df['Tipo Vinho']
                UvaVinho = base_df['Uva Vinho']
                ax.scatter(TipoVinho, UvaVinho, color='blue', alpha=0.6)
                plt.xticks(rotation=45)
                plt.title('Uvas de Vinos')
                plt.xlabel('Tipo Vinho')
                plt.ylabel('Uva Vinho')
                st.pyplot(fig)
                plt.close()

            with col2:
                st.subheader("Gráfico de Barras")
                fig, ax = plt.subplots(figsize=(8, 6))
                TipoVinho = base_df['Tipo Vinho']
                UvaVinho = base_df['Uva Vinho']
                ax.bar(TipoVinho, UvaVinho, color='skyblue')
                plt.xticks(rotation=45)
                plt.title('Comparación de Uvas de Vinos')
                plt.xlabel('Tipo Vinho')
                plt.ylabel('Uva Vinho')
                st.pyplot(fig)
                plt.close()    

        st.header("Visualizaciones con Seaborn")            
        with st.container():
            col1, col2 = st.columns(2)

            with col1:
                st.subheader("Gráfico de Violín")
                fig, ax = plt.subplots(figsize=(8, 6))
                sns.violinplot(data=dato_df, x ='Estilo Vinho',y='Característica Vinho 1')
                plt.xticks(rotation=45)
                plt.title('Distribución de Características por Estilo de Vino')
                st.pyplot(fig)
                plt.close()

            with col2:
                st.subheader("Gráfico de Caja")
                fig, ax = plt.subplots(figsize=(8, 6))
                sns.boxplot(data=dato_df, x ='Tipo de Queijo',y='Sabor Queijo 1')
                plt.xticks(rotation=45)
                plt.title('Distribución de Tipo de Queso por Sabor')
                st.pyplot(fig)
                plt.close()

        st.header("Visualizaciones Interactivas con Plotly")            
        with st.container():
            fig = px.line(base_df,
                          x='Tipo Vinho',
                          y='Uva Vinho',
                          title='Tendencia de Uva Vinho',
                          markers=True)
            st.plotly_chart(fig, use_container_width=True)

            fig = px.pie(dato_df,
                          values='Estilo Vinho',
                          names='Tipo Vinho',                          
                          title='Tendencia de Estilo de Vinho'                          )
            st.plotly_chart(fig, use_container_width=True)        

        st.header(" Sección Interactiva")    

        dataset_choice = st.radio(
            "Selecciona el conjunto de datos", 
            ["base_df","dato_df"]
        )

        if dataset_choice == "base_df":
            df = base_df
        else:
            df = dato_df

        chart_type = st.selectbox(
            "Selecciona el tipo de gráfico",
            ["Barras","Dispersión","Línea"]
        )

        x_axis = st.selectbox("Selecciona el eje X", df.columns)
        y_axis = st.selectbox("Selecciona el eje Y", df.columns)

        if chart_type == "Barras":
            fig = px.bar(df, x=x_axis, y=y_axis)
        elif chart_type == "Dispersión":
            fig = px.scatter(df,x=x_axis, y=y_axis)
        else:
            fig = px.line(df,x=x_axis, y=y_axis)

        st.plotly_chart(fig, use_container_width=True)


    except Exception as e:
        st.error(f" Error al cargar los datos: {str(e)}")    
        st.error("Por favor, verificar que los archivos existan en la carpeta 'datos' y tenga el formato correcto")
        
if __name__ == '__main__':
    graficos()

